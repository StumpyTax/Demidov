import csv
import datetime
import math
import re
import doctest
from os import SEEK_END
from typing import Any, Dict, Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import openpyxl.styles.numbers
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Border, Fill, Font, Side


class InputConnect:
  '''
  Обработка параметров вводимых пользователем.
  
  Attributes:
    name (str): Имя файла, полученное от пользователя.
    prof (str): Название профессии.
    vacancies_data (DataSet): Данные из файла.
  '''
  def __init__(self):
    '''
    Инициализирует объект InputConnect.

    Получает имя файла и проффессии от пользователя.
    '''
    self.name=input('Введите название файла: ')
    self.prof=input('Введите название профессии: ')
    self.vacancies_data=DataSet(self)

  def __create_bar_graph(self,fig,ax,data:dict,data_prof:dict,title:str,first_label:str,second_label:str)->None:
    '''
    Создает гистограмму из данных.

    Args:
        fig (_type_): Фигура
        ax (_type_): Оси
        data (dict): Данне по всем вакансиям
        data_prof (dict): Данные по профессии
        title (str): Заголовок гистограммы
        first_label (str): Заголовок для данных по всем вакансиям
        second_label (str): Заголовок для данных по профессии
    '''
    labels = list(data.keys())
    avg_sal = list(data.values())
    avg_prof_sal = list(data_prof.values())
    x = np.arange(len(labels)) 
    width = 0.35 

    rects1 = ax.bar(x - width/2, avg_sal, width, label=first_label)
    rects2 = ax.bar(x + width/2, avg_prof_sal, width, label=second_label)

    ax.set_title(title,fontsize = 8)
    for i in ax.get_yticklabels():
      i.set_fontsize(8)
    ax.set_xticks(x, labels,fontsize = 8, rotation= 90)

    ax.legend(fontsize = 8)
    ax.grid(visible=True,axis='y')
    fig.tight_layout()

  def __horizontal_bar_graph(self,fig,ax,data:dict,title:str)->None:
    '''
    Создает горизонтальную гистограмму из данных.

    Args:
        fig (_type_): Фигура
        ax (_type_): Оси
        data (dict): Данные
        title (str): Заголовок графика 
    '''
    labels = [i.replace(' ','\n') for i in list(data.keys())]
    y_pos = np.arange(len(labels))
    values=list(data.values())

    for i in ax.get_xticklabels():
      i.set_fontsize(8)

    ax.barh(y=y_pos,height=0.8,width=values,align='center')
    ax.set_yticks(y_pos, labels=labels,fontsize=6,horizontalalignment='right')
    ax.invert_yaxis()  # labels read top-to-bottom
    ax.set_title(title,fontsize=8)
    ax.grid(visible=True,axis='x')

  def __pie_graph(self,fig,ax,data:dict,title:str):
    '''
    Создает круговю диаграмму

    Args:
        fig (_type_): Фигура
        ax (_type_): Ось
        data (dict): Данные
        title (str): Заголовок диаграммы
    '''
    values=list(data.values())
    keys=list(data.keys())
    

    values.append(1-sum(values))
    keys.append('Другие')
    
    wedges, texts = ax.pie(x=values,labels=keys,textprops={'fontsize': 6})
    ax.axis('equal')
    ax.set_title(title,fontsize=8)

  def save_graph(self):
    '''
    Создает и сохраняет графики: две гистограммы, одну вертикальную гистограмму и одну круговю диаграмму.
    '''
    fig,ax=plt.subplots(2,2)
    self.__create_bar_graph(fig,ax[0,0],self.vacancies_data.salary_dynamic,
                                        self.vacancies_data.salary_dynamic_prof,
                                        'Уровень зарплаты по годам','средняя з/п',f'з/п {self.prof}')

    self.__create_bar_graph(fig,ax[0,1],self.vacancies_data.amount_dynamic,
                                        self.vacancies_data.amount_dynamic_prof,
                                        'Количество вакансий по годам','Кол-во вакансий',f'Кол-во вакансий {self.prof}')

    self.__horizontal_bar_graph(fig,ax[1,0],self.vacancies_data.get_top_10(self.vacancies_data.salary_dynamic_city),
                                            'Уровень зарплат по городам')

    self.__pie_graph(fig,ax[1,1],self.vacancies_data.get_top_10(self.vacancies_data.amount_dynamic_city),
                      'Доля вакансий по городам')

    plt.savefig('stat.png')


class Vacancy:
  '''
  Вакансия

  Attributes:
    name (str): Название вакансии.
    salary (Salary): Зарплата.
    area_name (str): Название города.
    date (datetime): Дата публикации вакансии.
  '''
  def __init__(self,v:dict)->None:
    '''
    Инициализирует объект Vacancy. Проводит первичную обработку данных

    Args:
        v (dict): Данные по вакансии
    >>> Vacancy({'name':'','salary_from':'1','salary_to':'8','salary_currency':'RUR','area_name':'Test1','published_at':'2007-12-03T19:15:55+0300' }).name
    ''
    >>> Vacancy({'name':'','salary_from':'0','salary_to':'150','salary_currency':'RUR','area_name':'Test1','published_at':'2007-12-03T19:15:55+0300' }).salary.salary_from
    '0'
    >>> Vacancy({'name':'','salary_from':'0','salary_to':'150','salary_currency':'RUR','area_name':'Test1','published_at':'2007-12-03T19:15:55+0300' }).area_name
    'Test1'
    '''
    vac=self.__clear(v)
    self.name=vac['name']
    self.salary=Salary(vac=vac)
    self.area_name=vac['area_name']
    self.date= datetime.datetime.strptime(vac['published_at'],"%Y-%m-%dT%H:%M:%S%z")

  def __clear(self,vac:dict)->dict:
    '''
    Очищает вакансию от лишних пробелов спец. символов и HTML тегов.

    Args:
        vac (dict): Вакансия

    Returns:
        dict: Очищенная вакансия
    '''
    n_dict={}
    for item in vac.items():
      l=item[1]
      l=re.sub(r"<[^>]+>", "",l,flags=re.S)
      l=re.sub(r" +",' ',l)
      l=re.sub(r"\u2002","",l)
      l=re.sub(r"\xa0"," ",l)
      l=l.strip()
      n_dict[item[0]]=l
    return n_dict

class DataSet:
  '''
  Данные из файла.

  Attributes:
    file_name (str): Название файла.
    vacancies_objects (List[Vacancy]): Список вакансий 
    salary_dynamic (dict): год:зарплата
    amount_dynamic (dict): год:кол-во вакансий
    salary_dynamic_prof (dict): год:зарплата у профессии
    amount_dynamic_prof (dict): год:кол-во вакансий на профессию
    salary_dynamic_city (dict): город: средняя зарплата зарплата
    amount_dynamic_city (dict): город: кол-во вакансий
  '''
  def __сsv_reader(self,ﬁle_name:str)->list[Vacancy]:
    '''
    Получает данные из файла.

    Args:
      file_name (str): Имя файла.

    Returns:
        list[Vacancy]: Список вакансий 
    '''
    with open(ﬁle_name,encoding="utf-8-sig") as f:
      reader = csv.DictReader(f)
      res=[]
      for row in reader:
        if (all(row.values()) and row.values().__len__()==reader.fieldnames.__len__()):
          vac=Vacancy(row)
          res.append(vac)

      if (f.seek(0, SEEK_END)==0):
        print('Пустой файл')
        exit()
      elif (res.__len__()==0):
        print('Нет данных')
        exit()
      return res

  def __init__(self,data:InputConnect) -> None:
    '''
    Инициализирует объект DataSet.

    Args:
        data (InputConnect): Данные полученные от пользователя.
    '''
    self.file_name =data.name

    self.vacancies_objects=self.__сsv_reader(data.name)
    self.salary_dynamic={}
    self.amount_dynamic={}
    self.salary_dynamic_prof={}
    self.amount_dynamic_prof={}
    self.salary_dynamic_city={}
    self.amount_dynamic_city={}

    self.collect_data(data.prof)


  def collect_data(self,prof:str)->None:
    '''
    Обрабатывает данные.

    Args:
        prof (str): Название профессии.
    '''
    city_vac_amount={}
    year_salary_summ={}
    prof_salary_sum={}

    for vac in self.vacancies_objects:
      if(vac.area_name in list(city_vac_amount.keys())):
        city_vac_amount[vac.area_name][0]+=vac.salary.one_cur()
        city_vac_amount[vac.area_name][1]+=1
      else:
        city_vac_amount[vac.area_name]=[vac.salary.one_cur(),1]

      if(vac.date.year in list(year_salary_summ.keys())):
        year_salary_summ[vac.date.year]+=vac.salary.one_cur()
        self.amount_dynamic[vac.date.year]+=1
      else:
        year_salary_summ[vac.date.year]=vac.salary.one_cur()
        self.amount_dynamic[vac.date.year]=1

      lower_name=str.lower(vac.name)
      if(str.lower(prof) in lower_name):
        if(vac.date.year in list(prof_salary_sum.keys())):
          prof_salary_sum[vac.date.year]+=vac.salary.one_cur()
          self.amount_dynamic_prof[vac.date.year]+=1
        else: 
          prof_salary_sum[vac.date.year]=vac.salary.one_cur()
          self.amount_dynamic_prof[vac.date.year]=1

    if(prof_salary_sum.items().__len__()>0):
      for profe in prof_salary_sum.items():
        self.salary_dynamic_prof[profe[0]]=int(math.trunc(profe[1]/self.amount_dynamic_prof[profe[0]]))
    else:
      for year in year_salary_summ.keys():
        self.salary_dynamic_prof[year]=0
        self.amount_dynamic_prof[year]=0
      
    for sum in year_salary_summ.items():
      self.salary_dynamic[sum[0]]=int(math.trunc(sum[1]/self.amount_dynamic[sum[0]]))
    
    for city in city_vac_amount.items():
      self.amount_dynamic_city[city[0]]=round(city[1][1]/self.vacancies_objects.__len__(),4)
      if(self.amount_dynamic_city[city[0]]>=0.01):
        self.salary_dynamic_city[city[0]]=int(math.trunc(city[1][0]/city_vac_amount[city[0]][1]))


  def get_top_10(self,dicti:dict[str,int])->dict[str,int]:
    '''
    Получает первые 10 наибольших значений словаря.

    Args:
        dicti (dict[str,int]): Словарь

    Returns:
        dict[str,int]: Топ 10 значений.
    '''
    res={}
    c=dicti.__len__() if dicti.__len__()<=10 else 10
    for i in range(c):
      max_value=-1
      good_vac={}
      for vac in dicti.items():
        if(vac[1]>max_value and not vac[0] in list(res.keys())):
          max_value=vac[1]
          good_vac=vac
      if( good_vac[1]>0.01):
        res[good_vac[0]]=good_vac[1]
    return res

  def print_data(self):
    '''
    Выводит статистику
    '''
    top_10_salary=self.get_top_10(self.salary_dynamic_city)
    top_10_amount=self.get_top_10(self.amount_dynamic_city)

    on_print=[f'Динамика уровня зарплат по годам: {self.salary_dynamic}',
              f'Динамика количества вакансий по годам: {self.amount_dynamic}',
              f'Динамика уровня зарплат по годам для выбранной профессии: {self.salary_dynamic_prof}',
              f'Динамика количества вакансий по годам для выбранной профессии: {self.amount_dynamic_prof}',
              f'Уровень зарплат по городам (в порядке убывания): {top_10_salary}',
              f'Доля вакансий по городам (в порядке убывания): {top_10_amount}']
    for i in on_print:
      print(i)
  
class Salary:
  '''
  Зарплата

  Attributes:
    salary_from (str):Минимальное значение зарплаты.
    salary_to (str):Максимальное значение зарплаты.
    salary_currency (str): Валюта зарплаты.
  '''
  def __init__(self,vac:dict) -> None:
    '''
    Инициализирует объект Salary

    Args:
        vac (dict): Вакансия
    '''
    self.salary_from=vac['salary_from']
    self.salary_to=vac['salary_to']
    self.salary_currency=vac['salary_currency']

  def avg(self)->float:
    '''
    Получает ср. значение зарплаты.

    Returns:
        float: Ср. значение зарплаты.
    >>> Salary({'salary_from':1,'salary_to':8,'salary_currency':'RUR'}).avg()
    4.5
    >>> Salary({'salary_from':0,'salary_to':0,'salary_currency':'RUR'}).avg()
    0.0
    >>> Salary({'salary_from':1000,'salary_to':100,'salary_currency':'RUR'}).avg()
    550.0
    '''
    return (float(self.salary_from)+float(self.salary_to))/2
      
  def one_cur(self)->float:
    '''
    Переводит ср. знач. зарплаты в рубли.

    Returns:
        float: ср. знач. зарплаты в рублях
    >>> Salary({'salary_from':100,'salary_to':1000,'salary_currency':'RUR'}).one_cur()
    550.0
    >>> Salary({'salary_from':100,'salary_to':1000,'salary_currency':'EUR'}).one_cur()
    32945.0
    >>> Salary({'salary_from':0,'salary_to':0,'salary_currency':'RUR'}).one_cur()
    0.0
    '''
    currency_to_rub = {
      "AZN": 35.68,  
      "BYR": 23.91,  
      "EUR": 59.90,  
      "GEL": 21.74,  
      "KGS": 0.76,  
      "KZT": 0.13,  
      "RUR": 1,  
      "UAH": 1.64,  
      "USD": 60.66,  
      "UZS": 0.0055,
      }
    return self.avg()*currency_to_rub[self.salary_currency]

class Report:
  '''
  Формирует отчет в формате pdf.

  Attributes:
    wb (Workbook):Excel файл.
    cities (WokrSheet): Лист статистики по городам в excel файле.
    years (WokrSheet): Лист статистики по годам в excel файле.
    data (DataSet): Данные по вакансиям.
    prof (str): Название профессии
  '''
  def __init__(self,data:InputConnect) -> None:
    '''
    Инициализирует объект Report. Формирует и создает pdf файл отчета.

    Args:
        data (InputConnect): Данные.
    '''
    self.wb=Workbook()
    self.cities=self.wb.create_sheet(title="Статистика по городам")
    self.years=self.wb.active
    self.wb.active.title="Статистика по годам"
    self.data=data.vacancies_data
    self.prof=data.prof
    self.Load_data(data=data)
    self.wb.save("report.xlsx")
    self.save_to_pdf()
  
  def __prep_yaer_stat_to_pdf(self)->list[list[str]]:
    '''
    Подготовка данных по годам для переноса в Excel.


    Returns:
        list[list[str]]: Данные по годам.
    '''
    res=[]
    for i,year in enumerate(self.data.salary_dynamic.keys()):
      res.append([year,self.data.salary_dynamic[year],self.data.salary_dynamic_prof[year],self.data.amount_dynamic[year],self.data.amount_dynamic_prof[year]])
    return res

  def __to_persent(self,dict:dict[str,int])->dict[str,str]:
    '''
    Переводит доли в проценты.

    Args:
        dict (dict[str,int]): Данные.

    Returns:
        dict[str,str]: Данные в процентах.
    '''
    res={}

    for i in dict.keys():
      res[i]=str(round(dict[i]*100,2))+'%'
    
    return res
  
  def save_to_pdf(self):
    '''
    Сохраняет отчет в виде pdf файла.
    '''
    env = Environment(loader=FileSystemLoader('.'))
    tp = env.get_template("pdf.html")    
    pdf_tp=tp.render({"image_file":'stat.png',
                      "prof":self.prof,
                      "data":self.__prep_yaer_stat_to_pdf(),
                      "salarySities":list(self.data.get_top_10(self.data.salary_dynamic_city).items()),
                      "amountCities":list(self.__to_persent(self.data.get_top_10(self.data.amount_dynamic_city)).items())})

    config = pdfkit.configuration(wkhtmltopdf='D:\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
    pdfkit.from_string(pdf_tp, 'report.pdf', configuration=config,options={"enable-local-file-access": ""})

  def Load_data(self,data:InputConnect)->None:
    '''
    Заполняет Excel файл данными.

    Args:
        data (InputConnect): Данные.
    '''
    self.years.append(['Год  ', 'Средняя зарплата', f'Средняя зарплата - {data.prof}','Количество вакансий', f'Количество вакансий - {data.prof}'])
    self.cities.append(['Город',	'Уровень зарплат ','','Город', 'Доля вакансий  '])

    self.__set_header_font(self.years)
    self.__set_header_font(self.cities)

    for row in data.vacancies_data.amount_dynamic.keys():
      self.years.append([row,data.vacancies_data.salary_dynamic[row],
                        data.vacancies_data.salary_dynamic_prof[row] if row in list(data.vacancies_data.salary_dynamic_prof.keys()) else "",
                         data.vacancies_data.amount_dynamic[row],
                         data.vacancies_data.amount_dynamic_prof[row]] if row in list(data.vacancies_data.amount_dynamic_prof.keys()) else "")

    cities_on_print=[]
    for row in data.vacancies_data.get_top_10(data.vacancies_data.salary_dynamic_city).items():
      cities_on_print.append([row[0],row[1],''])

    for i,row in enumerate(data.vacancies_data.get_top_10(data.vacancies_data.amount_dynamic_city).items()):
      if(i<cities_on_print.__len__()):
        cities_on_print[i].extend([row[0],float(row[1])])
      else:
        cities_on_print.append(['','','',row[0],float(row[1])])

    
    for city in cities_on_print:
      self.cities.append(city)

    self.__set_max_width(self.cities)
    self.__set_border(self.cities)
    self.__set_max_width(self.years)
    self.__set_border(self.years)

        
    for i in range(2,12):
      self.cities[f'E{i}'].number_format=openpyxl.styles.numbers.BUILTIN_FORMATS[10]

  def __set_header_font(self,sheet)->None:
    '''
    Стилизует шапку таблицы.

    Args:
        sheet (Worksheet): Таблица.
    '''
    for row in sheet.iter_rows(1):
      for cell in row:
        cell.font=Font(bold=True)

  def __set_border(self,sheet)->None:
    '''
    Стилизует границы таблицы

    Args:
        sheet (Worksheet): Таблица
    '''
    side=Side(border_style="thin", color="000000")
    for row in sheet.rows:
      for cell in row:
        if cell.value!='':
          cell.border=Border(right=side, bottom=side)
        else:
          cell.border=Border(right=side)

  def __set_max_width(self,sheet)->None:
    '''
    Устанавливает ширину заполненных ячеек таблицы.

    Args:
        sheet (Worksheet): Таблица
    '''
    ws = sheet
    dims = {}
    dims['C']=2
    for row in ws.columns:
        for cell in row:
            if cell.value:
              dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))+2))

    for col, value in dims.items():
        ws.column_dimensions[col].width = value

def get_stat():
  '''
  Формурет отчет.
  '''
  data=InputConnect()
  data.save_graph()
  rep=Report(data)


if(__name__=="__main__"):
  doctest.testmod()